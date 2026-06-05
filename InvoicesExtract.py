import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import json
import threading
import time
import sys
from typing import List

from openpyxl import Workbook, load_workbook
from google import genai
from google.genai import types
from pydantic import BaseModel, Field

# --- CẤU HÌNH ---
MODEL_NAME = "gemini-2.5-flash"
EXCEL_FILE = "invoice_extract.xlsx"
CONFIG_FILE = "config.json"

HEADERS = [
    "Tên File Gốc", "Số Hóa Đơn", "Ngày Hóa Đơn", "MST Người Bán", 
    "Tên Người Bán", "Tên Hàng Hóa", "Tiền Trước Thuế", 
    "Tiền Thuế", "Tổng Thanh Toán", "Tiền Tệ", "AI_Note"
]

# --- SCHEMA ---
class InvoiceInfo(BaseModel):
    invoice_no: str = Field(default="", description="Số hóa đơn")
    date: str = Field(default="", description="Ngày hóa đơn DD/MM/YYYY")
    seller_mst: str = Field(default="", description="Mã số thuế bên bán")
    seller_name: str = Field(default="", description="Tên công ty bán")
    items: List[str] = Field(default_factory=list, description="Danh sách TÊN hàng hóa/dịch vụ.")
    
    total_before_tax: float = Field(default=0.0, description="Tổng tiền trước thuế.")
    tax_amount: float = Field(default=0.0, description="Tiền thuế.")
    total_after_tax: float = Field(default=0.0, description="Tổng thanh toán.")
    currency: str = Field(default="VND", description="Loại tiền tệ (VD: VND, USD, EUR...)")

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"InvoicesExtract AI (anhnn@2025)")
        self.root.geometry("750x650")

        # --- CẤU HÌNH ICON GUI ---
        base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))

        icon_path = os.path.join(base_path, "Icon", "InvoicesExtract.ico")
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)
        # -------------------------
        
        # --- Flag điều khiển dừng ---
        self.stop_requested = False

        # GUI
        # --- GUI API Key (Đã thêm nút Hướng dẫn) ---
        frame_api = tk.Frame(root)
        frame_api.pack(pady=5)
        
        tk.Label(frame_api, text=f"Gemini API Key (Model: {MODEL_NAME}):").pack(side=tk.LEFT)
        self.btn_guide = tk.Button(frame_api, text="Hướng dẫn lấy Key", bg="#17a2b8", fg="white", font=("Arial", 9), command=self.show_api_guide)
        self.btn_guide.pack(side=tk.LEFT, padx=10)

        self.entry_api = tk.Entry(root, width=75, show="*")
        self.entry_api.pack(pady=5)
        # ------------------------------------------
        
        tk.Label(root, text="Thư mục chứa PDF/IMG/XML:").pack(pady=5)
        self.entry_folder = tk.Entry(root, width=75)
        self.entry_folder.pack(pady=5)
        self.btn_browse = tk.Button(root, text="Chọn Thư Mục", command=self.browse_folder)
        self.btn_browse.pack(pady=5)

        # Frame chứa 2 nút Bắt đầu và Dừng
        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=15)

        self.btn_start = tk.Button(btn_frame, text="BẮT ĐẦU XỬ LÝ", bg="#ffc107", fg="black", font=("Arial", 11, "bold"), width=20, command=self.start_thread)
        self.btn_start.pack(side=tk.LEFT, padx=10)
        
        self.btn_stop = tk.Button(btn_frame, text="DỪNG LẠI", bg="#dc3545", fg="white", font=("Arial", 11, "bold"), width=20, state=tk.DISABLED, command=self.request_stop)
        self.btn_stop.pack(side=tk.LEFT, padx=10)

        self.log_area = scrolledtext.ScrolledText(root, width=85, height=20)
        self.log_area.pack(pady=10)

        self.load_config()

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.entry_folder.delete(0, tk.END)
            self.entry_folder.insert(0, folder)

    def log(self, message):
        self.root.after(0, self._log_ui, message)

    def _log_ui(self, message):
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)

    def set_button_state(self, state, text):
        self.root.after(0, lambda: self.btn_start.config(state=state, text=text))

    def set_stop_button_state(self, state, text):
        self.root.after(0, lambda: self.btn_stop.config(state=state, text=text))

    def show_info(self, title, message):
        self.root.after(0, lambda: messagebox.showinfo(title, message))

    def show_api_guide(self):
        guide_text = """
HƯỚNG DẪN LẤY GEMINI API KEY MIỄN PHÍ

Bước 1: Mở trình duyệt web và truy cập vào trang:
         👉 https://aistudio.google.com/

Bước 2: Đăng nhập bằng tài khoản Google (Gmail) của bạn.

Bước 3: Nhìn sang menu bên trái, bấm vào "Get API key".

Bước 4: Bấm nút màu xanh "Create API key".
         - Nếu được hỏi, hãy chọn một Project có sẵn hoặc tạo mới.
         - Sao chép ⧉ (Copy) đoạn mã Key dài loằng ngoằng vừa hiện ra.

Bước 5: Quay lại ứng dụng này, dán (Paste) đoạn mã đó vào ô "Gemini API Key".

Lưu ý: API Key của bạn sẽ được lưu lại tự động cho những lần sử dụng sau. Tuyệt đối không chia sẻ Key này cho người lạ!
        """
        # Tạo một cửa sổ Toplevel để hiển thị nội dung dài
        guide_win = tk.Toplevel(self.root)
        guide_win.title("Hướng dẫn lấy Gemini API Key")
        guide_win.geometry("550x350")
        guide_win.resizable(False, False)
        
        # Đưa nội dung vào Text box dạng Read-Only để người dùng có thể copy link
        text_area = tk.Text(guide_win, wrap=tk.WORD, font=("Arial", 10), padx=15, pady=15, bg="#f8f9fa")
        text_area.insert(tk.END, guide_text.strip())
        text_area.config(state=tk.DISABLED) # Không cho phép sửa nội dung
        text_area.pack(expand=True, fill=tk.BOTH)

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    config = json.load(f)
                    self.entry_api.insert(0, config.get("api_key", ""))
                    self.entry_folder.insert(0, config.get("last_folder", ""))
            except: pass

    def save_config(self):
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump({"api_key": self.entry_api.get(), "last_folder": self.entry_folder.get()}, f)
        except: pass

    # --- Hàm xử lý sự kiện bấm nút Dừng ---
    def request_stop(self):
        if not self.stop_requested:
            self.stop_requested = True
            self.set_stop_button_state(tk.DISABLED, "Đang dừng...")
            self.log("\n[!] NHẬN LỆNH DỪNG: Sẽ dừng ngay sau khi xử lý xong file hiện tại...")

    def start_thread(self):
        self.save_config()
        self.stop_requested = False # Reset cờ mỗi khi chạy lại
        self.set_button_state(tk.DISABLED, "Đang chạy...")
        self.set_stop_button_state(tk.NORMAL, "DỪNG LẠI")
        thread = threading.Thread(target=self.process_invoices, daemon=True)
        thread.start()

    def fix_excel_headers(self, ws):
        current_headers = [cell.value for cell in ws[1]]
        if len(current_headers) < len(HEADERS) or "Tiền Tệ" not in current_headers:
            self.log("-> Chuẩn hóa Header Excel...")
            for col_num, header_title in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col_num, value=header_title)
            return True
        return False

    def append_to_excel(self, file_path, data_row):
        try:
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                if ws:
                    ws.append(HEADERS)
            else:
                wb = load_workbook(file_path)
                ws = wb.active
                if ws and self.fix_excel_headers(ws):
                    wb.save(file_path); wb = load_workbook(file_path); ws = wb.active
            
            if ws:
                while len(data_row) < len(HEADERS): data_row.append("")
                ws.append(data_row)
                
            wb.save(file_path)
            return True
        except PermissionError:
            return "Lỗi: File Excel đang mở, vui lòng đóng file!"
        except Exception as e:
            return str(e)

    def get_mime_type(self, filename):
        ext = filename.lower().split('.')[-1]
        if ext == "pdf": return "application/pdf"
        elif ext in ["jpg", "jpeg"]: return "image/jpeg"
        elif ext == "png": return "image/png"
        elif ext == "xml": return "text/plain" 
        return None

    def process_invoices(self):
        api_key = self.entry_api.get()
        folder_path = self.entry_folder.get()

        if not api_key or not folder_path:
            self.root.after(0, lambda: messagebox.showerror("Lỗi", "Vui lòng nhập đủ thông tin!"))
            self.set_button_state(tk.NORMAL, "BẮT ĐẦU XỬ LÝ")
            self.set_stop_button_state(tk.DISABLED, "DỪNG LẠI")
            return

        self.log(f"--- BẮT ĐẦU QUÉT ---")
        excel_path = os.path.join(folder_path, EXCEL_FILE)
        processed_files = set()

        if os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                if ws:
                    if self.fix_excel_headers(ws): wb.save(excel_path)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]: processed_files.add(row[0])
            except: pass

        try:
            valid_exts = (".pdf", ".jpg", ".jpeg", ".png", ".xml")
            all_files = [f for f in os.listdir(folder_path) if f.lower().endswith(valid_exts)]
        except: 
            self.set_button_state(tk.NORMAL, "BẮT ĐẦU XỬ LÝ")
            self.set_stop_button_state(tk.DISABLED, "DỪNG LẠI")
            return

        files_to_process = [f for f in all_files if f not in processed_files]
        if not files_to_process:
            self.log("-> Không có file mới.")
            self.set_button_state(tk.NORMAL, "BẮT ĐẦU XỬ LÝ")
            self.set_stop_button_state(tk.DISABLED, "DỪNG LẠI")
            return

        try:
            client = genai.Client(api_key=api_key)
        except Exception as e:
            self.log(f"Lỗi API: {e}")
            self.set_button_state(tk.NORMAL, "BẮT ĐẦU XỬ LÝ")
            self.set_stop_button_state(tk.DISABLED, "DỪNG LẠI")
            return

        count_ok = 0
        count_err = 0

        for filename in files_to_process:
            # --- KIỂM TRA CỜ DỪNG Ở ĐẦU VÒNG LẶP ---
            if self.stop_requested:
                self.log("\n[X] TIẾN TRÌNH ĐÃ DỪNG LẠI THEO YÊU CẦU CỦA NGƯỜI DÙNG.")
                break

            file_full_path = os.path.join(folder_path, filename)
            mime_type = self.get_mime_type(filename)
            
            if not mime_type:
                continue

            max_retries = 5 
            retry_count = 0
            success = False
            
            while retry_count < max_retries:
                try:
                    self.log(f">>> Đang xử lý: {filename}")
                    
                    with open(file_full_path, "rb") as f:
                        file_bytes = f.read()

                    prompt_text = """
                    Trích xuất dữ liệu hóa đơn/biên lai.
                    Lưu ý về Tiền tệ: Bắt buộc xác định loại tiền tệ (VND, USD, EUR...). 
                    Không bao gồm ký hiệu tiền tệ, dấu chấm hay dấu phẩy trong các trường số tiền.
                    """

                    response = client.models.generate_content(
                        model=MODEL_NAME,
                        contents=[
                            types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                            prompt_text
                        ],
                        config=types.GenerateContentConfig(
                            response_mime_type="application/json",
                            response_schema=InvoiceInfo
                        )
                    )
                    
                    if not response.text: raise ValueError("AI không phản hồi")
                    
                    data_dict = json.loads(response.text)
                    inv = InvoiceInfo(**data_dict) # type: ignore

                    if inv.currency.strip().upper() == "VND":
                        tbt = int(inv.total_before_tax)
                        tax = int(inv.tax_amount)
                        tat = int(inv.total_after_tax)
                    else:
                        tbt = float(inv.total_before_tax)
                        tax = float(inv.tax_amount)
                        tat = float(inv.total_after_tax)

                    items_str = "; ".join(inv.items) if inv.items else "Không xác định"

                    row_data = [
                        filename, inv.invoice_no, inv.date, inv.seller_mst, inv.seller_name,
                        items_str, tbt, tax, tat, inv.currency.upper(), "Hoàn thành"
                    ]
                    
                    res = self.append_to_excel(excel_path, row_data)
                    if res is True:
                        count_ok += 1
                        success = True
                        break 
                    else:
                        raise Exception(f"Lỗi ghi Excel: {res}")

                except Exception as e:
                    err_msg = str(e)
                    if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                        wait_time = (retry_count + 1) * 5 
                        self.log(f"!!! Quá tải (429). Đang chờ {wait_time}s rồi thử lại...")
                        
                        # --- Cắt nhỏ thời gian chờ để check lệnh dừng liên tục ---
                        for _ in range(wait_time):
                            if self.stop_requested:
                                break
                            time.sleep(1)
                            
                        if self.stop_requested:
                            break # Thoát vòng lặp while retry
                            
                        retry_count += 1
                        continue
                    else:
                        self.log(f"!!! Lỗi khác: {err_msg}")
                        row_data_err = [filename, "", "", "", "", "", 0, 0, 0, "", f"Lỗi: {err_msg}"]
                        self.append_to_excel(excel_path, row_data_err)
                        count_err += 1
                        success = True
                        break

            if not success and retry_count == max_retries and not self.stop_requested:
                self.log(f"!!! Bỏ qua file {filename} sau nhiều lần thử.")
                count_err += 1

        self.log(f"\n--- TỔNG KẾT ---")
        self.log(f"Thành công: {count_ok} | Lỗi: {count_err}")
        self.log(f"File kết quả: {excel_path}")
        
        msg = "Đã xử lý xong toàn bộ file!" if not self.stop_requested else "Đã dừng tiến trình!"
        self.show_info("Thông báo", msg)
        
        self.set_button_state(tk.NORMAL, "BẮT ĐẦU XỬ LÝ")
        self.set_stop_button_state(tk.DISABLED, "DỪNG LẠI")

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceApp(root)
    root.mainloop()